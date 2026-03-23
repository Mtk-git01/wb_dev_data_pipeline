
from __future__ import annotations

from pathlib import Path
from typing import List

import openpyxl
import pandas as pd

from src.config import AZE_BANKING_BULLETIN_XLSX_RAW_DIR
from src.extract_aze_bulletin_common import add_country_fields, dedup_keep_latest, find_sheet, iter_xlsx_files, parse_bulletin_period_from_filename, parse_row_period, to_num


def parse_aze_national_payment_systems_xlsx_file(xlsx_path: Path) -> pd.DataFrame:
    bulletin_period = parse_bulletin_period_from_filename(xlsx_path.name)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = find_sheet(wb, "4.1")

    rows = []
    current_year = None
    for r in range(12, ws.max_row + 1):
        token = ws.cell(r, 1).value
        if token is None:
            continue
        if str(token).strip().isdigit() and len(str(token).strip()) == 4:
            current_year = int(str(token).strip())
        period_date, period_type = parse_row_period(token, current_year)
        if period_date is None:
            continue

        rows.append({
            "period_date": period_date,
            "period_type": period_type,
            "rtgs_txn_count_thousand": to_num(ws.cell(r, 2).value),
            "rtgs_txn_amount_mn_azn": to_num(ws.cell(r, 3).value),
            "rtgs_avg_amount_thousand_azn": to_num(ws.cell(r, 4).value),
            "lvpcss_txn_count_thousand": to_num(ws.cell(r, 5).value),
            "lvpcss_txn_amount_mn_azn": to_num(ws.cell(r, 6).value),
            "lvpcss_avg_amount_azn": to_num(ws.cell(r, 7).value),
            "ips_txn_count_thousand": to_num(ws.cell(r, 8).value),
            "ips_txn_amount_mn_azn": to_num(ws.cell(r, 9).value),
            "source_file": xlsx_path.name,
            "bulletin_period": bulletin_period,
        })

    if not rows:
        raise ValueError(f"No rows parsed from sheet 4.1 in {xlsx_path.name}")
    return add_country_fields(pd.DataFrame(rows))


def load_aze_national_payment_systems_xlsx_raw() -> pd.DataFrame:
    frames: List[pd.DataFrame] = [parse_aze_national_payment_systems_xlsx_file(p) for p in iter_xlsx_files(AZE_BANKING_BULLETIN_XLSX_RAW_DIR)]
    return dedup_keep_latest(pd.concat(frames, ignore_index=True), ["country_iso", "period_date", "period_type"])
