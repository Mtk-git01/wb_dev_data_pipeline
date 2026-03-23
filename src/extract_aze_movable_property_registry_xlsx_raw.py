
from __future__ import annotations

from pathlib import Path
from typing import List

import openpyxl
import pandas as pd

from src.config import AZE_BANKING_BULLETIN_XLSX_RAW_DIR
from src.extract_aze_bulletin_common import add_country_fields, dedup_keep_latest, find_sheet, iter_xlsx_files, parse_bulletin_period_from_filename, parse_month_token, to_num


def parse_aze_movable_property_registry_xlsx_file(xlsx_path: Path) -> pd.DataFrame:
    bulletin_period = parse_bulletin_period_from_filename(xlsx_path.name)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = find_sheet(wb, "8")

    notice_year_cols = {}
    search_year_cols = {}
    for c in range(2, 9):
        if ws.cell(9, c).value:
            notice_year_cols[c] = int(ws.cell(9, c).value)
    for c in range(9, 16):
        if ws.cell(9, c).value:
            search_year_cols[c] = int(ws.cell(9, c).value)

    rows = []
    for r in range(10, ws.max_row + 1):
        month_num = parse_month_token(ws.cell(r, 1).value)
        if month_num is None:
            continue

        for c, year in notice_year_cols.items():
            value = to_num(ws.cell(r, c).value)
            if value is not None:
                rows.append({
                    "period_date": pd.Timestamp(year=year, month=month_num, day=1),
                    "period_type": "monthly",
                    "notices_entered_count": value,
                    "searches_count": None,
                    "source_file": xlsx_path.name,
                    "bulletin_period": bulletin_period,
                })

        for c, year in search_year_cols.items():
            value = to_num(ws.cell(r, c).value)
            if value is not None:
                rows.append({
                    "period_date": pd.Timestamp(year=year, month=month_num, day=1),
                    "period_type": "monthly",
                    "notices_entered_count": None,
                    "searches_count": value,
                    "source_file": xlsx_path.name,
                    "bulletin_period": bulletin_period,
                })

    if not rows:
        raise ValueError(f"No rows parsed from sheet 8 in {xlsx_path.name}")
    df = pd.DataFrame(rows)
    df = (
        df.groupby(["period_date", "period_type", "source_file", "bulletin_period"], as_index=False)[["notices_entered_count", "searches_count"]]
        .sum(min_count=1)
    )
    return add_country_fields(df)


def load_aze_movable_property_registry_xlsx_raw() -> pd.DataFrame:
    frames: List[pd.DataFrame] = [parse_aze_movable_property_registry_xlsx_file(p) for p in iter_xlsx_files(AZE_BANKING_BULLETIN_XLSX_RAW_DIR)]
    return dedup_keep_latest(pd.concat(frames, ignore_index=True), ["country_iso", "period_date", "period_type"])
