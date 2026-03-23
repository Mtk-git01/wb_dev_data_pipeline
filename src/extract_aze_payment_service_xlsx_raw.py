
from __future__ import annotations
from pathlib import Path
from typing import List
import openpyxl
import pandas as pd
from src.config import AZE_BANKING_BULLETIN_XLSX_RAW_DIR
from src.extract_aze_bulletin_common import add_country_fields, dedup_keep_latest, find_sheet, iter_xlsx_files, parse_bulletin_period_from_filename, parse_month_token, to_num

def parse_aze_payment_service_xlsx_file(xlsx_path: Path) -> pd.DataFrame:
    bulletin_period = parse_bulletin_period_from_filename(xlsx_path.name)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = find_sheet(wb, "4.3")
    rows = []
    current_year = None
    for r in range(12, ws.max_row + 1):
        token = ws.cell(r, 1).value
        if token is None:
            continue
        s = str(token).strip()
        if s.isdigit() and len(s) == 4:
            current_year = int(s)
            continue
        month_num = parse_month_token(token)
        if month_num is None or current_year is None:
            continue
        rows.append({
            "month": pd.Timestamp(year=current_year, month=month_num, day=1),
            "atm_total_units": to_num(ws.cell(r, 2).value),
            "atm_baku_units": to_num(ws.cell(r, 3).value),
            "atm_regions_units": to_num(ws.cell(r, 4).value),
            "pos_total_units": to_num(ws.cell(r, 5).value),
            "contactless_pos_total_units": to_num(ws.cell(r, 6).value),
            "retail_service_pos_total_units": to_num(ws.cell(r, 7).value),
            "retail_service_pos_baku_units": to_num(ws.cell(r, 8).value),
            "baku_pos_total_units": to_num(ws.cell(r, 9).value),
            "source_file": xlsx_path.name,
            "bulletin_period": bulletin_period,
        })
    if not rows:
        raise ValueError(f"No monthly rows parsed from sheet 4.3 in {xlsx_path.name}")
    return add_country_fields(pd.DataFrame(rows))

def load_aze_payment_service_xlsx_raw() -> pd.DataFrame:
    frames: List[pd.DataFrame] = [parse_aze_payment_service_xlsx_file(p) for p in iter_xlsx_files(AZE_BANKING_BULLETIN_XLSX_RAW_DIR)]
    return dedup_keep_latest(pd.concat(frames, ignore_index=True), ["country_iso", "month"])
