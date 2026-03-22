from __future__ import annotations

from pathlib import Path
import re
from typing import List

import openpyxl
import pandas as pd

from src.config import AZE_POLICY_BULLETIN_XLSX_RAW_DIR


def _find_sheet_31(workbook: openpyxl.Workbook):
    if "3.1 " in workbook.sheetnames:
        return workbook["3.1 "]
    if "3.1" in workbook.sheetnames:
        return workbook["3.1"]

    for name in workbook.sheetnames:
        if "3.1" in str(name):
            return workbook[name]

    raise ValueError("Sheet 3.1 not found")


def _normalize_label(x) -> str:
    if x is None:
        return ""
    return str(x).strip().lower()


def _parse_bulletin_period_from_filename(file_name: str) -> pd.Timestamp:
    """
    Expect filenames like statistical_bulletin_2026_01.xlsx
    """
    m = re.search(r"(\d{4})[_\-](\d{2})", file_name)
    if not m:
        raise ValueError(
            f"Could not parse bulletin period from filename: {file_name}. "
            "Use filenames like statistical_bulletin_2026_01.xlsx"
        )

    year = int(m.group(1))
    month = int(m.group(2))
    return pd.Timestamp(year=year, month=month, day=1)


def _find_row(ws, candidates: list[str]) -> int:
    """
    Search row labels in col A.
    """
    for r in range(1, ws.max_row + 1):
        label = _normalize_label(ws.cell(r, 1).value)
        if not label:
            continue

        for cand in candidates:
            if cand in label:
                return r

    raise ValueError(f"Target row not found for candidates: {candidates}")


def _build_month_grid(ws) -> list[tuple[int, pd.Timestamp]]:
    """
    Sheet 3.1 structure:
    - col A has year and month rows
    - policy rows are horizontal across columns? No:
      In the uploaded file, year/months are vertical in col A and the policy values
      are in fixed columns:
        col 13 -> corridor ceiling
        col 15 -> corridor floor
        col 17 -> refinancing rate
    So we parse the vertical year/month structure from col A.
    """
    rows = []
    current_year = None

    for r in range(14, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v is None:
            continue

        s = str(v).strip()

        # year row
        if re.fullmatch(r"\d{4}", s):
            current_year = int(s)
            continue

        # month row
        if current_year is not None and re.fullmatch(r"\d{1,2}", s):
            month = int(s)
            if 1 <= month <= 12:
                rows.append((r, pd.Timestamp(year=current_year, month=month, day=1)))

    if not rows:
        raise ValueError("No year/month grid parsed from sheet 3.1")

    return rows


def _to_num(x):
    if x is None:
        return None
    s = str(x).strip().replace(",", ".")
    if s in {"-", "—", ""}:
        return None
    try:
        return float(s)
    except Exception:
        return None


def parse_policy_bulletin_xlsx_file(xlsx_path: Path) -> pd.DataFrame:
    """
    Parse one CBAR Statistical Bulletin XLSX file from sheet 3.1.

    Output columns:
    - month
    - corridor_ceiling
    - corridor_floor
    - refinancing_rate
    - source_file
    - bulletin_period
    """
    bulletin_period = _parse_bulletin_period_from_filename(xlsx_path.name)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = _find_sheet_31(wb)

    month_rows = _build_month_grid(ws)

    # In the uploaded file:
    # col 13 -> Corridor ceiling, %
    # col 15 -> Corridor floor, %
    # col 17 -> Refinancing rate, %
    # Using fixed columns is the most stable for this workbook format.
    ceiling_col = 13
    floor_col = 15
    refi_col = 17

    records = []
    for row_idx, month in month_rows:
        records.append(
            {
                "month": month,
                "corridor_ceiling": _to_num(ws.cell(row_idx, ceiling_col).value),
                "corridor_floor": _to_num(ws.cell(row_idx, floor_col).value),
                "refinancing_rate": _to_num(ws.cell(row_idx, refi_col).value),
                "source_file": xlsx_path.name,
                "bulletin_period": bulletin_period,
            }
        )

    out = pd.DataFrame(records)
    out = out.dropna(subset=["month"]).reset_index(drop=True)
    return out


def load_aze_policy_bulletin_xlsx_raw() -> pd.DataFrame:
    raw_dir = Path(AZE_POLICY_BULLETIN_XLSX_RAW_DIR)

    if not raw_dir.exists():
        raise FileNotFoundError(f"Policy bulletin XLSX raw directory not found: {raw_dir}")

    xlsx_files = sorted(raw_dir.glob("*.xlsx"))
    if not xlsx_files:
        raise FileNotFoundError(f"No XLSX files found in {raw_dir}")

    frames: List[pd.DataFrame] = []
    for xlsx_file in xlsx_files:
        try:
            df = parse_policy_bulletin_xlsx_file(xlsx_file)
            frames.append(df)
        except Exception as e:
            print(f"[WARN] skipped {xlsx_file.name}: {e}")

    if not frames:
        raise ValueError("No policy bulletin XLSX files could be parsed")

    out = pd.concat(frames, ignore_index=True)

    # Dedup by month: keep latest bulletin_period
    out = (
        out.sort_values(["month", "bulletin_period", "source_file"])
        .drop_duplicates(subset=["month"], keep="last")
        .reset_index(drop=True)
    )

    return out