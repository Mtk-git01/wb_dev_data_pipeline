
from __future__ import annotations

from pathlib import Path
import re
from typing import Iterable, Optional

import openpyxl
import pandas as pd


AZ_MONTH_MAP = {
    "yanvar": 1,
    "fevral": 2,
    "mart": 3,
    "aprel": 4,
    "may": 5,
    "iyun": 6,
    "iyul": 7,
    "avqust": 8,
    "sentyabr": 9,
    "oktyabr": 10,
    "noyabr": 11,
    "dekabr": 12,
    "january": 1,
    "february": 2,
    "march": 3,
    "april": 4,
    "june": 6,
    "july": 7,
    "august": 8,
    "september": 9,
    "october": 10,
    "november": 11,
    "december": 12,
}

ROMAN_Q = {"i": 1, "ii": 2, "iii": 3, "iv": 4}


def normalize_text(x: object) -> str:
    if x is None:
        return ""
    s = str(x).strip().lower()
    s = s.replace("\n", " ")
    s = re.sub(r"\s+", " ", s)
    return s


def parse_bulletin_period_from_filename(file_name: str) -> pd.Timestamp:
    m = re.search(r"(\d{4})[_\-](\d{2})", file_name)
    if not m:
        raise ValueError(
            f"Could not parse bulletin period from filename: {file_name}. "
            "Expected filenames like statistical_bulletin_2026_01.xlsx"
        )
    return pd.Timestamp(year=int(m.group(1)), month=int(m.group(2)), day=1)


def find_sheet(workbook: openpyxl.Workbook, sheet_code: str):
    if sheet_code in workbook.sheetnames:
        return workbook[sheet_code]
    for name in workbook.sheetnames:
        if sheet_code in str(name):
            return workbook[name]
    raise ValueError(f"Sheet {sheet_code} not found. Available: {workbook.sheetnames}")


def iter_xlsx_files(raw_dir: str | Path) -> list[Path]:
    raw_path = Path(raw_dir)
    files = sorted(raw_path.glob("*.xlsx"))
    if not files:
        raise FileNotFoundError(f"No XLSX files found in {raw_path}")
    return files


def to_num(x: object) -> float | None:
    if x in [None, "", "-", "—"]:
        return None
    return pd.to_numeric(x, errors="coerce")


def is_year_token(x: object) -> bool:
    s = str(x).strip() if x is not None else ""
    return bool(re.fullmatch(r"\d{4}", s))


def parse_month_token(x: object) -> int | None:
    if x is None:
        return None
    s = normalize_text(x)
    if re.fullmatch(r"\d{1,2}", s):
        m = int(s)
        if 1 <= m <= 12:
            return m
    if s in AZ_MONTH_MAP:
        return AZ_MONTH_MAP[s]
    return None


def parse_row_period(token: object, current_year: int | None) -> tuple[pd.Timestamp | None, str | None]:
    if token is None:
        return None, None

    if isinstance(token, (pd.Timestamp, )):
        ts = pd.Timestamp(token)
        return ts.normalize(), "monthly"

    # openpyxl often returns datetime.datetime
    if hasattr(token, "year") and hasattr(token, "month") and hasattr(token, "day"):
        ts = pd.Timestamp(token)
        if ts.day == 1 and ts.month == 1:
            return ts.normalize(), "annual"
        return ts.normalize().replace(day=1), "monthly"

    s = str(token).strip()

    if re.fullmatch(r"\d{4}", s):
        return pd.Timestamp(year=int(s), month=1, day=1), "annual"

    m = parse_month_token(token)
    if m is not None and current_year is not None:
        return pd.Timestamp(year=current_year, month=m, day=1), "monthly"

    q_match = re.search(r"([ivx]{1,4}|q[1-4]|r[ivx]{1,4}|rq[1-4]|ri{0,3}v?)", normalize_text(s))
    year_match = re.search(r"(\d{4})", s)
    if year_match:
        year = int(year_match.group(1))
        q_num = None
        ns = normalize_text(s)
        if "q1" in ns or "qi" in ns or "ri" in ns:
            q_num = 1
        elif "q2" in ns or "qii" in ns or "rii" in ns:
            q_num = 2
        elif "q3" in ns or "qiii" in ns or "riii" in ns:
            q_num = 3
        elif "q4" in ns or "qiv" in ns or "riv" in ns:
            q_num = 4
        if q_num is not None:
            return pd.Timestamp(year=year, month=(q_num - 1) * 3 + 1, day=1), "quarterly"

    # 01.02.2005 style
    try:
        dt = pd.to_datetime(s, dayfirst=True, errors="raise")
        return pd.Timestamp(dt).replace(day=1), "monthly"
    except Exception:
        return None, None


def dedup_keep_latest(df: pd.DataFrame, key_cols: list[str]) -> pd.DataFrame:
    sort_cols = [c for c in key_cols + ["bulletin_period", "source_file"] if c in df.columns]
    if not sort_cols:
        return df.reset_index(drop=True)
    return (
        df.sort_values(sort_cols)
        .drop_duplicates(subset=key_cols, keep="last")
        .reset_index(drop=True)
    )


def add_country_fields(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["country_name"] = "Azerbaijan"
    out["country_iso"] = "AZE"
    return out


def safe_merge_periodic(left: pd.DataFrame, right: pd.DataFrame, how: str = "outer") -> pd.DataFrame:
    join_cols = ["country_name", "country_iso", "period_date", "period_type"]
    if left.empty:
        return right.copy()
    if right.empty:
        return left.copy()
    return left.merge(right, on=join_cols, how=how)
