
from __future__ import annotations

from pathlib import Path
from typing import List

import openpyxl
import pandas as pd

from src.config import AZE_BANKING_BULLETIN_XLSX_RAW_DIR
from src.extract_aze_bulletin_common import add_country_fields, dedup_keep_latest, find_sheet, iter_xlsx_files, parse_bulletin_period_from_filename, to_num


def parse_aze_business_portfolio_xlsx_file(xlsx_path: Path) -> pd.DataFrame:
    bulletin_period = parse_bulletin_period_from_filename(xlsx_path.name)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = find_sheet(wb, "5.7")

    date_cols = {}
    for c in range(2, ws.max_column + 1):
        v = ws.cell(6, c).value
        if v is None:
            continue
        ts = pd.to_datetime(v, errors="coerce")
        if pd.notna(ts):
            date_cols[c] = pd.Timestamp(ts).replace(day=1)

    row_map = {
        7: "business_loans_total_mn_azn",
        9: "large_business_loans_mn_azn",
        10: "medium_business_loans_mn_azn",
        11: "small_business_loans_mn_azn",
        12: "micro_business_loans_mn_azn",
    }

    rows = []
    for c, period_date in date_cols.items():
        rec = {
            "period_date": period_date,
            "period_type": "monthly",
            "source_file": xlsx_path.name,
            "bulletin_period": bulletin_period,
        }
        for r, name in row_map.items():
            rec[name] = to_num(ws.cell(r, c).value)
        rows.append(rec)

    if not rows:
        raise ValueError(f"No rows parsed from sheet 5.7 in {xlsx_path.name}")
    return add_country_fields(pd.DataFrame(rows))


def load_aze_business_portfolio_xlsx_raw() -> pd.DataFrame:
    frames: List[pd.DataFrame] = [parse_aze_business_portfolio_xlsx_file(p) for p in iter_xlsx_files(AZE_BANKING_BULLETIN_XLSX_RAW_DIR)]
    return dedup_keep_latest(pd.concat(frames, ignore_index=True), ["country_iso", "period_date", "period_type"])
