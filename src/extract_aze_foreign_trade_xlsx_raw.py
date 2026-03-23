
from __future__ import annotations

from pathlib import Path
from typing import List

import openpyxl
import pandas as pd

from src.config import AZE_BANKING_BULLETIN_XLSX_RAW_DIR
from src.extract_aze_bulletin_common import add_country_fields, dedup_keep_latest, find_sheet, iter_xlsx_files, parse_bulletin_period_from_filename, parse_month_token, to_num


def parse_aze_foreign_trade_xlsx_file(xlsx_path: Path) -> pd.DataFrame:
    bulletin_period = parse_bulletin_period_from_filename(xlsx_path.name)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = find_sheet(wb, "1.5")

    rows = []
    current_year = None
    q_map = {"i rüb": 1, "ii rüb": 2, "iii rüb": 3, "iv rüb": 4}
    for r in range(19, ws.max_row + 1):
        token = ws.cell(r, 1).value
        if token is None:
            continue
        s = str(token).strip().lower()
        if s.isdigit() and len(s) == 4:
            current_year = int(s)
            period_date = pd.Timestamp(year=current_year, month=1, day=1)
            period_type = "annual"
        elif any(k in s for k in q_map) and current_year is not None:
            q = 1 if "i  rüb" in s or "i rüb" in s and "ii" not in s and "iii" not in s and "iv" not in s else None
            if "ii" in s:
                q = 2
            elif "iii" in s:
                q = 3
            elif "iv" in s:
                q = 4
            if q is None:
                continue
            period_date = pd.Timestamp(year=current_year, month=(q - 1) * 3 + 1, day=1)
            period_type = "quarterly"
        else:
            continue

        rows.append({
            "period_date": period_date,
            "period_type": period_type,
            "total_exports_ths_usd": to_num(ws.cell(r, 2).value),
            "exports_yoy_pct": to_num(ws.cell(r, 3).value),
            "exports_non_cis_ths_usd": to_num(ws.cell(r, 4).value),
            "exports_non_cis_yoy_pct": to_num(ws.cell(r, 5).value),
            "exports_cis_ths_usd": to_num(ws.cell(r, 6).value),
            "exports_cis_yoy_pct": to_num(ws.cell(r, 7).value),
            "total_imports_ths_usd": to_num(ws.cell(r, 8).value),
            "imports_yoy_pct": to_num(ws.cell(r, 9).value),
            "source_file": xlsx_path.name,
            "bulletin_period": bulletin_period,
        })

    if not rows:
        raise ValueError(f"No rows parsed from sheet 1.5 in {xlsx_path.name}")
    return add_country_fields(pd.DataFrame(rows))


def load_aze_foreign_trade_xlsx_raw() -> pd.DataFrame:
    frames: List[pd.DataFrame] = [parse_aze_foreign_trade_xlsx_file(p) for p in iter_xlsx_files(AZE_BANKING_BULLETIN_XLSX_RAW_DIR)]
    return dedup_keep_latest(pd.concat(frames, ignore_index=True), ["country_iso", "period_date", "period_type"])
