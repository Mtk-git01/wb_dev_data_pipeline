
from __future__ import annotations

from pathlib import Path
from typing import List

import openpyxl
import pandas as pd

from src.config import AZE_BANKING_BULLETIN_XLSX_RAW_DIR
from src.extract_aze_bulletin_common import add_country_fields, dedup_keep_latest, find_sheet, iter_xlsx_files, normalize_text, parse_bulletin_period_from_filename, to_num


def parse_aze_interest_rates_xlsx_file(xlsx_path: Path) -> pd.DataFrame:
    bulletin_period = parse_bulletin_period_from_filename(xlsx_path.name)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = find_sheet(wb, "3.2")

    rows = []
    current_date = None
    current_rec = None

    for r in range(13, ws.max_row + 1):
        token = ws.cell(r, 2).value
        if token is None:
            continue

        token_norm = normalize_text(token)
        parsed_date = pd.to_datetime(token, dayfirst=True, errors="coerce")
        if pd.notna(parsed_date):
            if current_rec:
                rows.append(current_rec)
            current_date = pd.Timestamp(parsed_date).replace(day=1)
            current_rec = {
                "period_date": current_date,
                "period_type": "monthly",
                "source_file": xlsx_path.name,
                "bulletin_period": bulletin_period,
            }
            continue

        if current_rec is None:
            continue

        if "manat" in token_norm:
            current_rec["deposit_avg_rate_azn_pct"] = to_num(ws.cell(r, 3).value)
            current_rec["deposit_legal_rate_azn_pct"] = to_num(ws.cell(r, 4).value)
            current_rec["deposit_individual_rate_azn_pct"] = to_num(ws.cell(r, 5).value)
            current_rec["loan_avg_rate_azn_pct"] = to_num(ws.cell(r, 13).value)
            current_rec["loan_legal_rate_azn_pct"] = to_num(ws.cell(r, 14).value)
            current_rec["loan_individual_rate_azn_pct"] = to_num(ws.cell(r, 15).value)
        elif "xarici valyuta" in token_norm or "foreign" in token_norm:
            current_rec["deposit_avg_rate_fx_pct"] = to_num(ws.cell(r, 3).value)
            current_rec["deposit_legal_rate_fx_pct"] = to_num(ws.cell(r, 4).value)
            current_rec["deposit_individual_rate_fx_pct"] = to_num(ws.cell(r, 5).value)
            current_rec["loan_avg_rate_fx_pct"] = to_num(ws.cell(r, 13).value)
            current_rec["loan_legal_rate_fx_pct"] = to_num(ws.cell(r, 14).value)
            current_rec["loan_individual_rate_fx_pct"] = to_num(ws.cell(r, 15).value)

    if current_rec:
        rows.append(current_rec)

    if not rows:
        raise ValueError(f"No rows parsed from sheet 3.2 in {xlsx_path.name}")
    return add_country_fields(pd.DataFrame(rows))


def load_aze_interest_rates_xlsx_raw() -> pd.DataFrame:
    frames: List[pd.DataFrame] = [parse_aze_interest_rates_xlsx_file(p) for p in iter_xlsx_files(AZE_BANKING_BULLETIN_XLSX_RAW_DIR)]
    return dedup_keep_latest(pd.concat(frames, ignore_index=True), ["country_iso", "period_date", "period_type"])
