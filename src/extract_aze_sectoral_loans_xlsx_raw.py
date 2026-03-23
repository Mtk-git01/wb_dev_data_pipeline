
from __future__ import annotations

from pathlib import Path
from typing import List

import openpyxl
import pandas as pd

from src.config import AZE_BANKING_BULLETIN_XLSX_RAW_DIR
from src.extract_aze_bulletin_common import add_country_fields, dedup_keep_latest, find_sheet, iter_xlsx_files, parse_bulletin_period_from_filename, parse_row_period, to_num


def parse_aze_sectoral_loans_xlsx_file(xlsx_path: Path) -> pd.DataFrame:
    bulletin_period = parse_bulletin_period_from_filename(xlsx_path.name)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = find_sheet(wb, "2.8")

    rows = []
    current_year = None
    for r in range(15, ws.max_row + 1):
        token = ws.cell(r, 1).value
        period_date, period_type = parse_row_period(token, current_year)
        if isinstance(token, (int, float)) and str(int(token)).isdigit() and len(str(int(token))) == 4:
            current_year = int(token)
        if period_date is None:
            continue

        rec = {
            "period_date": period_date,
            "period_type": period_type,
            "real_sector_loans_total_mn_azn": to_num(ws.cell(r, 2).value),
            "real_sector_overdue_loans_mn_azn": to_num(ws.cell(r, 3).value),
            "real_sector_overdue_share_pct": to_num(ws.cell(r, 4).value),
            "trade_and_services_loans_mn_azn": to_num(ws.cell(r, 5).value),
            "mining_utilities_loans_mn_azn": to_num(ws.cell(r, 7).value),
            "agriculture_loans_mn_azn": to_num(ws.cell(r, 9).value),
            "construction_loans_mn_azn": to_num(ws.cell(r, 11).value),
            "industry_manufacturing_loans_mn_azn": to_num(ws.cell(r, 13).value),
            "transport_communication_loans_mn_azn": to_num(ws.cell(r, 15).value),
            "households_loans_mn_azn": to_num(ws.cell(r, 17).value),
            "real_estate_mortgage_loans_mn_azn": to_num(ws.cell(r, 19).value),
            "other_sectors_loans_mn_azn": to_num(ws.cell(r, 25).value),
            "source_file": xlsx_path.name,
            "bulletin_period": bulletin_period,
        }
        rows.append(rec)

    if not rows:
        raise ValueError(f"No rows parsed from sheet 2.8 in {xlsx_path.name}")
    return add_country_fields(pd.DataFrame(rows))


def load_aze_sectoral_loans_xlsx_raw() -> pd.DataFrame:
    frames: List[pd.DataFrame] = [parse_aze_sectoral_loans_xlsx_file(p) for p in iter_xlsx_files(AZE_BANKING_BULLETIN_XLSX_RAW_DIR)]
    return dedup_keep_latest(pd.concat(frames, ignore_index=True), ["country_iso", "period_date", "period_type"])
