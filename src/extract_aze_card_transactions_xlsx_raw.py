
from __future__ import annotations
from pathlib import Path
from typing import List
import openpyxl
import pandas as pd
from src.config import AZE_BANKING_BULLETIN_XLSX_RAW_DIR
from src.extract_aze_bulletin_common import add_country_fields, dedup_keep_latest, find_sheet, iter_xlsx_files, parse_bulletin_period_from_filename, parse_month_token, to_num

def parse_aze_card_transactions_xlsx_file(xlsx_path: Path) -> pd.DataFrame:
    bulletin_period = parse_bulletin_period_from_filename(xlsx_path.name)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = find_sheet(wb, "4.5")
    rows = []
    current_year = None
    for r in range(16, ws.max_row + 1):
        token = ws.cell(r, 1).value
        if token is None:
            continue
        s = str(token).strip()
        if s.isdigit() and len(s) == 4:
            current_year = int(s)
            continue
        month_num = parse_month_token(token)
        if month_num is None or current_year is None:
            continue
        rows.append({
            "month": pd.Timestamp(year=current_year, month=month_num, day=1),
            "payment_cards_total_thousand": to_num(ws.cell(r, 2).value),
            "payment_cards_contactless_thousand": to_num(ws.cell(r, 3).value),
            "debit_social_cards_thousand": to_num(ws.cell(r, 4).value),
            "debit_salary_cards_thousand": to_num(ws.cell(r, 5).value),
            "debit_other_cards_thousand": to_num(ws.cell(r, 6).value),
            "credit_cards_thousand": to_num(ws.cell(r, 7).value),
            "card_txn_count_thousand": to_num(ws.cell(r, 8).value),
            "card_txn_amount_mn_azn": to_num(ws.cell(r, 9).value),
            "cash_withdrawal_atm_count_thousand": to_num(ws.cell(r, 10).value),
            "cash_withdrawal_atm_amount_mn_azn": to_num(ws.cell(r, 11).value),
            "cash_withdrawal_pos_count_thousand": to_num(ws.cell(r, 12).value),
            "cash_withdrawal_pos_amount_mn_azn": to_num(ws.cell(r, 13).value),
            "non_cash_atm_count_thousand": to_num(ws.cell(r, 14).value),
            "non_cash_atm_amount_mn_azn": to_num(ws.cell(r, 15).value),
            "non_cash_pos_count_thousand": to_num(ws.cell(r, 16).value),
            "non_cash_pos_amount_mn_azn": to_num(ws.cell(r, 17).value),
            "contactless_pos_count_thousand": to_num(ws.cell(r, 18).value),
            "contactless_pos_amount_mn_azn": to_num(ws.cell(r, 19).value),
            "ecommerce_count_thousand": to_num(ws.cell(r, 20).value),
            "ecommerce_amount_mn_azn": to_num(ws.cell(r, 21).value),
            "self_service_terminal_count_thousand": to_num(ws.cell(r, 22).value),
            "self_service_terminal_amount_mn_azn": to_num(ws.cell(r, 23).value),
            "source_file": xlsx_path.name,
            "bulletin_period": bulletin_period,
        })
    if not rows:
        raise ValueError(f"No monthly rows parsed from sheet 4.5 in {xlsx_path.name}")
    return add_country_fields(pd.DataFrame(rows))

def load_aze_card_transactions_xlsx_raw() -> pd.DataFrame:
    frames: List[pd.DataFrame] = [parse_aze_card_transactions_xlsx_file(p) for p in iter_xlsx_files(AZE_BANKING_BULLETIN_XLSX_RAW_DIR)]
    return dedup_keep_latest(pd.concat(frames, ignore_index=True), ["country_iso", "month"])
