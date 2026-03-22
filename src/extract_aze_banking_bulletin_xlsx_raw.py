from __future__ import annotations

from pathlib import Path
import re
from typing import List

import openpyxl
import pandas as pd

from src.config import AZE_BANKING_BULLETIN_XLSX_RAW_DIR


def _find_sheet_52(workbook: openpyxl.Workbook):
    """
    Prefer exact sheet name '5.2', otherwise use any sheet containing '5.2'.
    """
    if "5.2" in workbook.sheetnames:
        return workbook["5.2"]

    for name in workbook.sheetnames:
        if "5.2" in str(name):
            return workbook[name]

    raise ValueError("Sheet 5.2 not found")


def _normalize_label(x) -> str:
    if x is None:
        return ""
    return str(x).strip().lower()


def _parse_bulletin_period_from_filename(file_name: str) -> pd.Timestamp:
    """
    Expect filenames like:
    - statistical_bulletin_2024_03.xlsx
    - statistical_bulletin_2025_12.xlsx
    """
    m = re.search(r"(\d{4})[_\-](\d{2})", file_name)
    if not m:
        raise ValueError(
            f"Could not parse bulletin period from filename: {file_name}. "
            "Use filenames like statistical_bulletin_2024_03.xlsx"
        )

    year = int(m.group(1))
    month = int(m.group(2))
    return pd.Timestamp(year=year, month=month, day=1)


def _find_row(ws, candidates: list[str]) -> int:
    """
    Find row index where col A contains one of the target labels.
    """
    for r in range(1, ws.max_row + 1):
        label = _normalize_label(ws.cell(r, 1).value)
        if not label:
            continue

        for cand in candidates:
            if cand in label:
                return r

    raise ValueError(f"Target row not found for candidates: {candidates}")


def _extract_month_columns(ws) -> list[tuple[int, pd.Timestamp]]:
    """
    In sheet 5.2:
    - row 6 contains dates
    - row 7 contains 'Cəmi' / 'Xarici valyutada'
    Keep only 'Cəmi' / 'Total' columns.
    """
    result = []

    header_row_dates = 6
    header_row_types = 7

    for col in range(2, ws.max_column + 1):
        date_val = ws.cell(header_row_dates, col).value
        type_val = ws.cell(header_row_types, col).value

        if date_val is None or type_val is None:
            continue

        type_text = str(type_val).strip().lower()
        if type_text not in {"cəmi", "total"}:
            continue

        dt = pd.to_datetime(date_val, errors="coerce")
        if pd.isna(dt):
            continue

        month = dt.to_period("M").to_timestamp()
        result.append((col, month))

    if not result:
        raise ValueError("No month columns parsed from sheet 5.2")

    return result


def _extract_series_from_row(
    ws,
    row_idx: int,
    month_cols: list[tuple[int, pd.Timestamp]],
) -> list[float]:
    vals = []
    for col, _month in month_cols:
        val = ws.cell(row_idx, col).value
        vals.append(pd.to_numeric(val, errors="coerce"))
    return vals


def parse_banking_bulletin_xlsx_file(xlsx_path: Path) -> pd.DataFrame:
    """
    Parse one CBAR Statistical Bulletin XLSX file from sheet 5.2.

    Output columns:
    - month
    - bank_total_assets_mn_azn
    - bank_loans_customers_mn_azn
    - bank_deposits_total_mn_azn
    - source_file
    - bulletin_period
    """
    bulletin_period = _parse_bulletin_period_from_filename(xlsx_path.name)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = _find_sheet_52(wb)

    month_cols = _extract_month_columns(ws)

    total_assets_row = _find_row(
        ws,
        [
            "11. cəmi aktivlər",
            "11. total assets",
        ],
    )

    loans_row = _find_row(
        ws,
        [
            "7. müştərilərə verilən kreditlər",
            "7. loans to customers",
        ],
    )

    deposits_row = _find_row(
        ws,
        [
            "1. depozitlər (maliyyə institutları istisna olmaqla)",
            "1. deposits (excluding financial institutions)",
        ],
    )

    assets_vals = _extract_series_from_row(ws, total_assets_row, month_cols)
    loans_vals = _extract_series_from_row(ws, loans_row, month_cols)
    deposits_vals = _extract_series_from_row(ws, deposits_row, month_cols)

    out = pd.DataFrame(
        {
            "month": [m for _, m in month_cols],
            "bank_total_assets_mn_azn": assets_vals,
            "bank_loans_customers_mn_azn": loans_vals,
            "bank_deposits_total_mn_azn": deposits_vals,
            "source_file": xlsx_path.name,
            "bulletin_period": bulletin_period,
        }
    )

    out = out.dropna(subset=["month"]).copy()
    return out.reset_index(drop=True)


def load_aze_banking_bulletin_xlsx_raw() -> pd.DataFrame:
    raw_dir = Path(AZE_BANKING_BULLETIN_XLSX_RAW_DIR)

    if not raw_dir.exists():
        raise FileNotFoundError(f"Banking bulletin XLSX raw directory not found: {raw_dir}")

    xlsx_files = sorted(raw_dir.glob("*.xlsx"))
    if not xlsx_files:
        raise FileNotFoundError(f"No XLSX files found in {raw_dir}")

    frames: List[pd.DataFrame] = []
    for xlsx_file in xlsx_files:
        try:
            df = parse_banking_bulletin_xlsx_file(xlsx_file)
            frames.append(df)
        except Exception as e:
            print(f"[WARN] skipped {xlsx_file.name}: {e}")

    if not frames:
        raise ValueError("No banking bulletin XLSX files could be parsed")

    out = pd.concat(frames, ignore_index=True)

    # Dedup rule:
    # If the same month appears in multiple bulletin files,
    # keep the row from the latest bulletin_period.
    out = (
        out.sort_values(["month", "bulletin_period", "source_file"])
        .drop_duplicates(subset=["month"], keep="last")
        .reset_index(drop=True)
    )

    return out