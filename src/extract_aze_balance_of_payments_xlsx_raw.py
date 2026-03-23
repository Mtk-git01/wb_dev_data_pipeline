
from __future__ import annotations

from pathlib import Path
from typing import List

import openpyxl
import pandas as pd

from src.config import AZE_BANKING_BULLETIN_XLSX_RAW_DIR
from src.extract_aze_bulletin_common import add_country_fields, dedup_keep_latest, find_sheet, iter_xlsx_files, normalize_text, parse_bulletin_period_from_filename, parse_row_period, to_num


ROW_MATCHES = {
    "current_account_balance_mn_usd": ["cari hesab", "current account"],
    "trade_balance_mn_usd": ["xarici ticarət balans", "foreign trade balance"],
    "goods_exports_total_mn_usd": ["malların ixracı", "goods export"],
    "oil_gas_exports_mn_usd": ["neft-qaz sektoru", "oil-gas sector"],
    "other_exports_mn_usd": ["digər sektorlar", "other sectors"],
    "goods_imports_total_mn_usd": ["malların idxalı", "goods import"],
    "oil_gas_imports_mn_usd": ["neft-qaz sektoru", "oil-gas sector"],
    "other_imports_mn_usd": ["digər sektorlar", "other sectors"],
}


def parse_aze_balance_of_payments_xlsx_file(xlsx_path: Path) -> pd.DataFrame:
    bulletin_period = parse_bulletin_period_from_filename(xlsx_path.name)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = find_sheet(wb, "1.4")

    periods = {}
    for c in range(2, ws.max_column + 1):
        period_date, period_type = parse_row_period(ws.cell(6, c).value, None)
        if period_date is not None:
            periods[c] = (period_date, period_type)

    rows = []
    for c, (period_date, period_type) in periods.items():
        rec = {
            "period_date": period_date,
            "period_type": period_type,
            "source_file": xlsx_path.name,
            "bulletin_period": bulletin_period,
        }
        current_label_bucket = None
        for r in range(7, 60):
            label = normalize_text(ws.cell(r, 1).value)
            if not label:
                continue
            # disambiguate repeated "oil-gas sector" / "other sectors" rows using prior top-level row
            if "malların ixracı" in label or "goods export" in label:
                current_label_bucket = "exports"
            elif "malların idxalı" in label or "goods import" in label:
                current_label_bucket = "imports"

            value = to_num(ws.cell(r, c).value)
            if value is None:
                continue

            if "cari hesab" in label or "current account" in label:
                rec["current_account_balance_mn_usd"] = value
            elif "xarici ticarət balans" in label or "foreign trade balance" in label:
                rec["trade_balance_mn_usd"] = value
            elif "malların ixracı" in label or "goods export" in label:
                rec["goods_exports_total_mn_usd"] = value
            elif "malların idxalı" in label or "goods import" in label:
                rec["goods_imports_total_mn_usd"] = value
            elif ("neft-qaz sektoru" in label or "oil-gas sector" in label) and current_label_bucket == "exports":
                rec["oil_gas_exports_mn_usd"] = value
            elif ("digər sektorlar" in label or "other sectors" in label) and current_label_bucket == "exports":
                rec["other_exports_mn_usd"] = value
            elif ("neft-qaz sektoru" in label or "oil-gas sector" in label) and current_label_bucket == "imports":
                rec["oil_gas_imports_mn_usd"] = value
            elif ("digər sektorlar" in label or "other sectors" in label) and current_label_bucket == "imports":
                rec["other_imports_mn_usd"] = value

        rows.append(rec)

    if not rows:
        raise ValueError(f"No rows parsed from sheet 1.4 in {xlsx_path.name}")
    return add_country_fields(pd.DataFrame(rows))


def load_aze_balance_of_payments_xlsx_raw() -> pd.DataFrame:
    frames: List[pd.DataFrame] = [parse_aze_balance_of_payments_xlsx_file(p) for p in iter_xlsx_files(AZE_BANKING_BULLETIN_XLSX_RAW_DIR)]
    return dedup_keep_latest(pd.concat(frames, ignore_index=True), ["country_iso", "period_date", "period_type"])
