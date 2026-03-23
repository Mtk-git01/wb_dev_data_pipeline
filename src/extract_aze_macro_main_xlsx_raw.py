
from __future__ import annotations

from pathlib import Path
from typing import List

import openpyxl
import pandas as pd

from src.config import AZE_BANKING_BULLETIN_XLSX_RAW_DIR
from src.extract_aze_bulletin_common import add_country_fields, dedup_keep_latest, find_sheet, iter_xlsx_files, parse_bulletin_period_from_filename, parse_month_token, to_num


def parse_aze_macro_main_xlsx_file(xlsx_path: Path) -> pd.DataFrame:
    bulletin_period = parse_bulletin_period_from_filename(xlsx_path.name)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = find_sheet(wb, "1.1")

    rows = []
    current_year = None
    for r in range(13, ws.max_row + 1):
        token = ws.cell(r, 1).value
        if token is None:
            continue
        s = str(token).strip()
        if s.isdigit() and len(s) == 4:
            period_date = pd.Timestamp(year=int(s), month=1, day=1)
            period_type = "annual"
            current_year = int(s)
        else:
            month_num = parse_month_token(token)
            if month_num is None or current_year is None:
                continue
            period_date = pd.Timestamp(year=current_year, month=month_num, day=1)
            period_type = "monthly"

        rows.append({
            "period_date": period_date,
            "period_type": period_type,
            "nominal_gdp_mn_azn": to_num(ws.cell(r, 2).value),
            "real_gdp_growth_pct": to_num(ws.cell(r, 3).value),
            "gdp_deflator_pct": to_num(ws.cell(r, 4).value),
            "non_oil_gdp_mn_azn": to_num(ws.cell(r, 5).value),
            "non_oil_gdp_growth_pct": to_num(ws.cell(r, 6).value),
            "capital_investment_mn_azn": to_num(ws.cell(r, 7).value),
            "capital_investment_growth_pct": to_num(ws.cell(r, 8).value),
            "nominal_income_mn_azn": to_num(ws.cell(r, 9).value),
            "nominal_income_growth_pct": to_num(ws.cell(r, 10).value),
            "avg_monthly_wage_azn": to_num(ws.cell(r, 11).value),
            "avg_monthly_wage_growth_pct": to_num(ws.cell(r, 12).value),
            "cpi_monthly_pct": to_num(ws.cell(r, 13).value),
            "cpi_12m_pct": to_num(ws.cell(r, 14).value),
            "cpi_annual_avg_pct": to_num(ws.cell(r, 15).value),
            "source_file": xlsx_path.name,
            "bulletin_period": bulletin_period,
        })

    if not rows:
        raise ValueError(f"No rows parsed from sheet 1.1 in {xlsx_path.name}")
    return add_country_fields(pd.DataFrame(rows))


def load_aze_macro_main_xlsx_raw() -> pd.DataFrame:
    frames: List[pd.DataFrame] = [parse_aze_macro_main_xlsx_file(p) for p in iter_xlsx_files(AZE_BANKING_BULLETIN_XLSX_RAW_DIR)]
    return dedup_keep_latest(pd.concat(frames, ignore_index=True), ["country_iso", "period_date", "period_type"])
