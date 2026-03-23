from __future__ import annotations

from pathlib import Path
from typing import List
import re
import unicodedata

import openpyxl
import pandas as pd

from src.config import AZE_BANKING_BULLETIN_XLSX_RAW_DIR
from src.extract_aze_bulletin_common import (
    add_country_fields,
    dedup_keep_latest,
    find_sheet,
    iter_xlsx_files,
    normalize_text,
    parse_bulletin_period_from_filename,
    to_num,
)


def normalize_match_text(s) -> str:
    """
    regexp:

    e.g.,:
      QİK -> qik
      qi̇k -> qik
    """
    if s is None:
        return ""

    s = str(s)
    s = normalize_text(s)

    # Unicode
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))

    # lower
    s = s.lower()

    # -
    s = (
        s.replace("–", "-")
         .replace("—", "-")
         .replace("−", "-")
         .replace("-", "-")
    )

    # blank
    s = s.replace("\xa0", " ")

    # reg
    s = re.sub(r"\s*/\s*", "/", s)
    s = re.sub(r"\s+", " ", s).strip()

    return s


def parse_pct_value(value):
    """
    
    4.5% -> 4.5
    0.045 -> 4.5
    4.5 -> 4.5
    """
    num = to_num(value)
    if num is None:
        return None

    # adjust (e.g.,0.045 -> 4.5)
    return num * 100 if num <= 1 else num


def build_npl_label_map() -> dict:
    """
    正規化後ラベル -> 出力カラム名 の対応。
    """
    return {
        "qeyri-islək kredit (qik)": "npl_total_mn_azn",
        "qeyri-islek kredit (qik)": "npl_total_mn_azn",
        "qik/kredit portfeli": "npl_ratio_pct",

        "- biznes kreditləri": "npl_business_mn_azn",
        "- biznes kreditleri": "npl_business_mn_azn",
        "biznes kreditləri": "npl_business_mn_azn",
        "biznes kreditleri": "npl_business_mn_azn",

        "- istehlak kreditləri": "npl_consumer_mn_azn",
        "- istehlak kreditleri": "npl_consumer_mn_azn",
        "istehlak kreditləri": "npl_consumer_mn_azn",
        "istehlak kreditleri": "npl_consumer_mn_azn",

        "- ipoteka kreditləri": "npl_mortgage_mn_azn",
        "- ipoteka kreditleri": "npl_mortgage_mn_azn",
        "ipoteka kreditləri": "npl_mortgage_mn_azn",
        "ipoteka kreditleri": "npl_mortgage_mn_azn",

        "- biznes (qik)/biznes kreditləri": "npl_business_ratio_pct",
        "- biznes (qik)/biznes kreditleri": "npl_business_ratio_pct",
        "biznes (qik)/biznes kreditləri": "npl_business_ratio_pct",
        "biznes (qik)/biznes kreditleri": "npl_business_ratio_pct",

        "- istehlak (qik)/istehlak kreditləri": "npl_consumer_ratio_pct",
        "- istehlak (qik)/istehlak kreditleri": "npl_consumer_ratio_pct",
        "istehlak (qik)/istehlak kreditləri": "npl_consumer_ratio_pct",
        "istehlak (qik)/istehlak kreditleri": "npl_consumer_ratio_pct",

        "- ipoteka (qik)/ipoteka kreditləri": "npl_mortgage_ratio_pct",
        "- ipoteka (qik)/ipoteka kreditleri": "npl_mortgage_ratio_pct",
        "ipoteka (qik)/ipoteka kreditləri": "npl_mortgage_ratio_pct",
        "ipoteka (qik)/ipoteka kreditleri": "npl_mortgage_ratio_pct",
    }


def is_pct_metric(col_name: str) -> bool:
    return col_name.endswith("_pct")


def parse_aze_npl_structure_xlsx_file(xlsx_path: Path) -> pd.DataFrame:
    bulletin_period = parse_bulletin_period_from_filename(xlsx_path.name)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = find_sheet(wb, "5.6")

    label_map = build_npl_label_map()

    # date column
    date_cols = {}
    for c in range(2, ws.max_column + 1):
        v = ws.cell(6, c).value
        ts = pd.to_datetime(v, errors="coerce")
        if pd.notna(ts):
            date_cols[c] = pd.Timestamp(ts).replace(day=1)

    rows = []

    # sheet 5.6 
    for c, period_date in date_cols.items():
        rec = {
            "period_date": period_date,
            "period_type": "monthly",
            "source_file": xlsx_path.name,
            "bulletin_period": bulletin_period,
        }

        for r in range(7, min(ws.max_row, 40) + 1):
            raw_label = ws.cell(r, 1).value
            label = normalize_match_text(raw_label)
            value = ws.cell(r, c).value

            if not label:
                continue

            out_col = label_map.get(label)
            if not out_col:
                continue

            if is_pct_metric(out_col):
                parsed_value = parse_pct_value(value)
            else:
                parsed_value = to_num(value)

            if parsed_value is None:
                continue

            rec[out_col] = parsed_value

        rows.append(rec)

    if not rows:
        raise ValueError(f"No rows parsed from sheet 5.6 in {xlsx_path.name}")

    out = pd.DataFrame(rows)

    # Check main columns
    expected_any = [
        "npl_total_mn_azn",
        "npl_ratio_pct",
        "npl_business_mn_azn",
        "npl_consumer_mn_azn",
        "npl_mortgage_mn_azn",
    ]
    if not any(col in out.columns for col in expected_any):
        raise ValueError(
            f"Sheet 5.6 parsed but no expected NPL metrics were captured in {xlsx_path.name}"
        )

    return add_country_fields(out)


def load_aze_npl_structure_xlsx_raw() -> pd.DataFrame:
    frames: List[pd.DataFrame] = [
        parse_aze_npl_structure_xlsx_file(p)
        for p in iter_xlsx_files(AZE_BANKING_BULLETIN_XLSX_RAW_DIR)
    ]

    out = pd.concat(frames, ignore_index=True)

    return dedup_keep_latest(
        out,
        ["country_iso", "period_date", "period_type"],
    )